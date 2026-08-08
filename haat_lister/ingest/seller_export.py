"""§4.1 -- the operator's own seller-panel export, as listing rows.

THE HIGHEST-YIELD ROUTE, and the reason is arithmetic rather than taste. A
fetch of a marketplace product page gives us a title, a description written for
that marketplace, and whatever photographs survive nine predicates. An export
from the panel the operator already signs into gives us the SKU, the real price,
the real weight, the variant list and the manufacturer's own image URLs -- for
every product at once, with no request to anybody.

It is also the only route where the data is unambiguously theirs.

WHAT THIS MODULE DOES NOT DO. It does not decide anything about a row. Mapped
columns become `FieldValue`s with `FieldSource.OPERATOR`, and everything after
that -- enrichment, the Tier-1 image chain, policy screening, the provenance
gate -- is `pipeline.process_page`'s job, exactly as for a fetched page. An
import that scored its own rows would be a second pipeline, and the first thing
it would skip is the gate.

COLUMNS ARE NEVER SILENTLY DROPPED. Anything unmapped is reported back, by name,
with a sample value. A seller export is somebody's inventory and the columns we
do not understand are the ones most likely to matter to them.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from hashlib import blake2s
from pathlib import Path
from typing import TYPE_CHECKING

import yaml

from ..models import Confidence, FieldSource, FieldValue, ProductRecord, Provenance
from ..utils.logging import get_logger

if TYPE_CHECKING:
    from ..config import Settings

log = get_logger(__name__)

MAX_BYTES = 16 * 1024 * 1024
MAX_ROWS = 10_000
SUFFIXES = (".csv", ".tsv", ".txt", ".xlsx", ".xlsm")


class ExportError(Exception):
    """Bad input from a human. Always a message they can act on."""


# ---------------------------------------------------------------------------
# What we can accept a column for
# ---------------------------------------------------------------------------
#
# Deliberately NOT the full 19-column header. These are the fields an export can
# state as fact; everything else on the row is derived, and a derived field that
# arrived by import would be a guess wearing an operator's authority.
#
# `gi_region` is absent, and its absence is load-bearing (§7). There is no
# column name, no alias and no fuzzy match that can write it.

TARGETS: dict[str, tuple[str, ...]] = {
    "source_url": ("url", "product url", "link", "product link", "listing url", "page"),
    "title": ("title", "name", "product name", "product title", "item name"),
    "description": ("description", "details", "long description", "product description"),
    "price_inr": ("price", "mrp", "selling price", "price inr", "rate", "amount"),
    "weight_g": ("weight", "weight g", "weight grams", "shipping weight", "net weight"),
    "hs_code": ("hs code", "hsn", "hsn code", "tariff code", "commodity code"),
    "category_slug": ("category", "product category", "department", "type"),
    "subcategory_slug": ("subcategory", "sub category", "sub type"),
    "length_cm": ("length", "length cm", "depth"),
    "width_cm": ("width", "width cm", "breadth"),
    "height_cm": ("height", "height cm"),
    "availability": ("availability", "stock status", "in stock", "status"),
    "stock_qty": ("quantity", "stock", "stock qty", "qty", "inventory"),
    "sizes": ("size", "sizes", "size chart", "available sizes"),
    "seller_note": ("note", "seller note", "remarks", "comments"),
    "image_urls": (
        "image", "images", "image url", "image urls", "photo", "photos",
        "main image", "image src", "additional images", "image link",
    ),
}

# Columns a seller export genuinely has and this tool has nowhere to put. Named
# so the unmapped report can say "recognised, but there is no column for it"
# rather than leaving the operator wondering whether the mapper is broken.
KNOWN_UNUSED: dict[str, tuple[str, ...]] = {
    "sku": ("sku", "seller sku", "item code", "style code", "product code"),
    "brand": ("brand", "manufacturer", "label", "vendor"),
    "material": ("material", "fabric", "composition"),
    "colour": ("colour", "color", "shade"),
}

# Never accepted from a file, under any header. Listed rather than merely
# omitted so that the refusal is visible at the point somebody adds a column.
REFUSED_TARGETS: frozenset[str] = frozenset({"gi_region"})

_SPLIT = re.compile(r"[|,;\n]+")


@dataclass
class Column:
    """One column of the file, and what we think it is."""

    index: int
    header: str
    samples: list[str] = field(default_factory=list)
    target: str = ""
    confidence: float = 0.0

    @property
    def mapped(self) -> bool:
        return bool(self.target)


@dataclass
class Export:
    """A parsed export, before anything has been decided about its rows."""

    path: Path
    columns: list[Column]
    rows: list[list[str]]
    signature: str
    profile_used: str = ""

    @property
    def unmapped(self) -> list[Column]:
        """Shown to the operator, never discarded quietly."""
        return [c for c in self.columns if not c.mapped]

    @property
    def mapping(self) -> dict[str, int]:
        return {c.target: c.index for c in self.columns if c.mapped}

    def value(self, row: list[str], target: str) -> str:
        index = self.mapping.get(target)
        if index is None or index >= len(row):
            return ""
        return row[index].strip()


# ---------------------------------------------------------------------------
# Reading the file
# ---------------------------------------------------------------------------


def _read_tabular(path: Path) -> list[list[str]]:
    raw = path.read_bytes()
    if len(raw) > MAX_BYTES:
        raise ExportError(
            f"{path.name} is {len(raw) / 1_048_576:.0f} MB, over the "
            f"{MAX_BYTES // 1_048_576} MB limit."
        )
    text = raw.decode("utf-8-sig", "replace")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t;|")
    except csv.Error:
        # One column, or a file the sniffer cannot read. Comma is the honest
        # default and a wrong guess is visible immediately in the mapper.
        dialect = csv.excel
    rows = csv.reader(io.StringIO(text), dialect)
    return [row for row in rows if any(cell.strip() for cell in row)]


def _read_xlsx(path: Path) -> list[list[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover -- depends on the extra
        raise ExportError(
            "Reading .xlsx needs the spreadsheet extra: pip install 'haat-lister[sheets]'. "
            "Or save the sheet as CSV, which this reads with no extra installed."
        ) from exc

    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book.active
    if sheet is None:  # pragma: no cover -- empty workbook
        raise ExportError(f"{path.name} has no sheets.")
    rows: list[list[str]] = []
    for raw in sheet.iter_rows(values_only=True):
        cells = ["" if cell is None else str(cell).strip() for cell in raw]
        if any(cells):
            rows.append(cells)
        if len(rows) > MAX_ROWS + 1:
            break
    book.close()
    return rows


def read(path: Path) -> tuple[list[str], list[list[str]]]:
    """Header row and data rows. Raises `ExportError` for anything unreadable."""
    if not path.is_file():
        raise ExportError(f"{path} is not a file.")
    if path.suffix.lower() not in SUFFIXES:
        raise ExportError(
            f"{path.name} is not a spreadsheet. Export your catalogue as .csv or .xlsx."
        )

    rows = _read_xlsx(path) if path.suffix.lower() in (".xlsx", ".xlsm") else _read_tabular(path)
    if not rows:
        raise ExportError(f"{path.name} is empty.")
    if len(rows) - 1 > MAX_ROWS:
        raise ExportError(
            f"{path.name} has {len(rows) - 1:,} rows, over the {MAX_ROWS:,} limit. "
            "Split it and run the parts as separate jobs."
        )

    header = [cell.strip() for cell in rows[0]]
    if not any(header):
        raise ExportError(f"{path.name} has no header row.")
    return header, rows[1:]


# ---------------------------------------------------------------------------
# Guessing what each column is
# ---------------------------------------------------------------------------


def _normalise(header: str) -> str:
    return re.sub(r"[^a-z0-9 ]+", " ", header.lower()).strip()


def _score(header: str, alias: str) -> float:
    """How much this header looks like that alias.

    Exact and containment beat similarity, and the ratio is a tiebreak rather
    than the mechanism. Fuzzy matching alone maps `price` to `price_inr` and
    also maps `priced at` and `pricelist name`, and a mapper that is confidently
    wrong is worse than one that asks.
    """
    clean = _normalise(header)
    if not clean:
        return 0.0
    if clean == alias:
        return 1.0
    words = clean.split()
    if alias in words:
        return 0.92
    if clean.startswith(alias + " ") or clean.endswith(" " + alias):
        return 0.88
    if alias in clean:
        return 0.8
    ratio = SequenceMatcher(None, clean, alias).ratio()
    return ratio if ratio >= 0.86 else 0.0


def auto_map(columns: list[Column]) -> None:
    """Fill in `target` where a column is recognisable. Never guesses twice.

    Each target is claimed by at most one column -- the best-scoring one -- so
    a sheet with `Price` and `Price (USD)` does not silently take whichever came
    last. The loser is reported as unmapped, which is the honest outcome: two
    columns both look like the price and a human should say which.
    """
    claims: dict[str, tuple[float, Column]] = {}
    for column in columns:
        best_target, best_score = "", 0.0
        for target, aliases in TARGETS.items():
            if target in REFUSED_TARGETS:  # pragma: no cover -- belt and braces
                continue
            score = max(_score(column.header, alias) for alias in aliases)
            if score > best_score:
                best_target, best_score = target, score
        if best_score < 0.8:
            continue
        held = claims.get(best_target)
        if held is None or best_score > held[0]:
            if held is not None:
                held[1].target, held[1].confidence = "", 0.0
            claims[best_target] = (best_score, column)
            column.target, column.confidence = best_target, best_score


def signature(header: list[str]) -> str:
    """A stable key for "an export shaped like this one".

    Order-insensitive and case-insensitive, because a panel that adds a column
    or reorders two should still recognise the profile the operator saved. It is
    a convenience key, not a security boundary -- a collision costs a wrong
    default in a mapper the operator is looking at.
    """
    parts = sorted(_normalise(h) for h in header if h.strip())
    return blake2s(" ".join(parts).encode("utf-8"), digest_size=8).hexdigest()


# ---------------------------------------------------------------------------
# Profiles
# ---------------------------------------------------------------------------


def profiles_dir(settings: Settings) -> Path:
    return settings.root / "profiles"


def save_profile(settings: Settings, name: str, export: Export) -> Path:
    """Remember this mapping for the next export with the same headers.

    Stored as YAML next to the run rather than in the ledger because it is the
    operator's, it is worth reading, and it is worth putting in version control
    alongside whatever else describes their catalogue.
    """
    safe = re.sub(r"[^a-z0-9_-]+", "-", name.lower()).strip("-") or "export"
    folder = profiles_dir(settings)
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{safe}.yaml"
    path.write_text(
        yaml.safe_dump(
            {
                "name": safe,
                "signature": export.signature,
                "headers": [c.header for c in export.columns],
                "mapping": {c.header: c.target for c in export.columns if c.mapped},
            },
            sort_keys=False,
            allow_unicode=True,
        ),
        encoding="utf-8",
    )
    log.info("Saved column profile %s", path)
    return path


def load_profiles(settings: Settings) -> list[dict]:
    folder = profiles_dir(settings)
    if not folder.is_dir():
        return []
    found: list[dict] = []
    for path in sorted(folder.glob("*.yaml")):
        try:
            data = yaml.safe_load(path.read_text(encoding="utf-8"))
        except (OSError, yaml.YAMLError) as exc:
            log.warning("Ignoring unreadable profile %s: %s", path.name, exc)
            continue
        if isinstance(data, dict) and data.get("mapping"):
            found.append(data)
    return found


def apply_profile(export: Export, profile: dict) -> None:
    """Overwrite the auto-map with a saved one, by header name.

    Refused targets are dropped on the way in. A profile is a file an operator
    can edit, so `gi_region: seller_region` is a thing somebody will eventually
    write, and it must not work.
    """
    mapping = {
        _normalise(k): v
        for k, v in profile.get("mapping", {}).items()
        if v in TARGETS and v not in REFUSED_TARGETS
    }
    if not mapping:
        return
    for column in export.columns:
        column.target = mapping.get(_normalise(column.header), "")
        column.confidence = 1.0 if column.target else 0.0
    export.profile_used = str(profile.get("name", ""))


# ---------------------------------------------------------------------------
# The one entry point
# ---------------------------------------------------------------------------


def parse(path: Path, settings: Settings | None = None) -> Export:
    """Read an export and propose a mapping. Nothing is committed here."""
    header, rows = read(path)
    columns = [
        Column(
            index=index,
            header=name,
            samples=[r[index].strip() for r in rows[:3] if index < len(r) and r[index].strip()],
        )
        for index, name in enumerate(header)
    ]
    auto_map(columns)
    export = Export(path=path, columns=columns, rows=rows, signature=signature(header))

    if settings is not None:
        for profile in load_profiles(settings):
            if profile.get("signature") == export.signature:
                apply_profile(export, profile)
                log.info("Applied saved profile %s", export.profile_used)
                break
    return export


# ---------------------------------------------------------------------------
# Rows
# ---------------------------------------------------------------------------


def _number(text: str) -> float | None:
    cleaned = re.sub(r"[^\d.]", "", text.replace(",", ""))
    try:
        return float(cleaned) if cleaned else None
    except ValueError:  # pragma: no cover -- the regex admits only digits and dots
        return None


def _weight_grams(text: str) -> int | None:
    """Grams, from whatever unit the panel wrote.

    Silence beats a guess here: an unrecognised unit leaves the field empty and
    the row reaches review, which is `§7`'s "money and customs fields are never
    silently guessed" applied to the thing customs actually charges on.
    """
    value = _number(text)
    if value is None:
        return None
    lowered = text.lower()
    if "kg" in lowered or "kilo" in lowered:
        return int(value * 1000)
    if "lb" in lowered or "pound" in lowered:
        return int(value * 453.592)
    if "oz" in lowered or "ounce" in lowered:
        return int(value * 28.3495)
    return int(value)


def to_record(
    export: Export,
    row: list[str],
    provenance: Provenance,
    settings: Settings,
) -> ProductRecord:
    """One spreadsheet row as a record, with nothing decided about it yet.

    Every value lands as `FieldSource.OPERATOR` at HIGH confidence, because that
    is what it is: a human's own catalogue, not an inference off a page. The
    fields it does not carry stay absent, so `missing_required` reports them and
    the row reaches review rather than shipping with a plausible blank.
    """
    from ..models import FetchStage
    from ..pipeline import new_record

    url = export.value(row, "source_url")
    record = new_record(url, provenance, settings.identity)
    record.fetch_stage = FetchStage.SELLER_EXPORT

    def stated(target: str) -> None:
        if text := export.value(row, target):
            setattr(record, target, FieldValue.found(text, FieldSource.OPERATOR, Confidence.HIGH))

    for target in (
        "title", "description", "category_slug", "subcategory_slug",
        "hs_code", "availability", "sizes", "seller_note",
    ):
        stated(target)

    if (price := _number(export.value(row, "price_inr"))) is not None:
        record.price_inr = FieldValue.found(int(price), FieldSource.OPERATOR, Confidence.HIGH)
    if (grams := _weight_grams(export.value(row, "weight_g"))) is not None:
        record.weight_g = FieldValue.found(grams, FieldSource.OPERATOR, Confidence.HIGH)
    for axis in ("length_cm", "width_cm", "height_cm"):
        if (cm := _number(export.value(row, axis))) is not None:
            setattr(record, axis, FieldValue.found(int(cm), FieldSource.OPERATOR, Confidence.HIGH))
    if (qty := _number(export.value(row, "stock_qty"))) is not None:
        record.stock_qty = FieldValue.found(int(qty), FieldSource.OPERATOR, Confidence.HIGH)

    # Image URLs go into the ordinary candidate list, so they meet the identical
    # nine predicates. An operator's own export is not evidence that a URL is
    # reachable, decodable, big enough, or servable to a stranger.
    record.image_candidates = [
        part.strip()
        for part in _SPLIT.split(export.value(row, "image_urls"))
        if part.strip().startswith(("http://", "https://"))
    ][: settings.config.images.max_images_per_product]

    if export.unmapped:
        # §4.1. Never silently discarded. A seller export is somebody's whole
        # inventory and the columns we did not understand are exactly the ones
        # most likely to matter to them.
        record.note(
            "Columns this import did not use: "
            + ", ".join(c.header for c in export.unmapped[:8])
            + ("..." if len(export.unmapped) > 8 else "")
            + ". Map them on the import screen if any of them matter."
        )
    return record


def known_unused(header: str) -> str:
    """Which recognisable-but-unwritable field this column is, if any.

    `sku` and `brand` are real and useful and there is no haat column for them.
    Saying "recognised, no column for it" is a different message to "not
    recognised", and only one of them tells the operator to stop looking.
    """
    for name, aliases in KNOWN_UNUSED.items():
        if any(_score(header, alias) >= 0.8 for alias in aliases):
            return name
    return ""
